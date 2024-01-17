from pathlib import Path
from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtWidgets import QApplication, QWidget, QLabel, QLineEdit, QPushButton, QFileDialog
import sys

class MinhaApp(QWidget):
    def __init__(self):
        super().__init__()

        self.initUI()

    def initUI(self):
        self.setWindowTitle('Minha Aplicação')
        self.setGeometry(100, 100, 400, 200)

        self.label_path = QLabel('Caminho do Arquivo:', self)
        self.label_path.move(10, 10)

        self.entry_path = QLineEdit(self)
        self.entry_path.setGeometry(10, 30, 300, 20)

        self.button_browse = QPushButton('Selecionar Arquivo', self)
        self.button_browse.setGeometry(10, 60, 150, 30)
        self.button_browse.clicked.connect(self.selecionar_arquivo)

        self.button_processar = QPushButton('Processar Arquivo', self)
        self.button_processar.setGeometry(180, 60, 150, 30)
        self.button_processar.clicked.connect(self.processar_arquivo)

        self.show()

    def selecionar_arquivo(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Selecionar Arquivo')
        if file_path:
            self.entry_path.setText(file_path)

    def processar_arquivo(self):
        file_path, _ = QFileDialog.getOpenFileName(self, 'Selecionar Arquivo para Processar')

        if not file_path:
            return  # Se o caminho do arquivo não estiver definido, não faça nada

        workbook = load_workbook(file_path)

        sheet_name = "Codigos Car"
        worksheet: Worksheet = workbook[sheet_name]

        # Encontre o índice da última coluna com valores
        last_column_index = max(worksheet.max_column, 1)  # Certifique-se de que o índice não seja inferior a 1

        # Adicione uma nova coluna após a última coluna com valores
        worksheet.insert_cols(last_column_index + 1)

        list_names = []

        row: tuple[Cell]
        for row in worksheet.iter_rows():
            line = ""
            for cell in row:
                if cell.value is not None:
                    words = cell.value.split()
                    first_chars = [word[0] for word in words]
                    line += "".join(first_chars)
            if line:
                list_names.append(line)

        # Adicione os valores nas células da nova coluna (supondo que list_names contenha os dados)
        for index, value in enumerate(list_names, start=3):
            worksheet.cell(row=index, column=last_column_index + 1, value=value)

        workbook.save(file_path)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    minha_app = MinhaApp()
    sys.exit(app.exec())
