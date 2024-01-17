from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FILE = Path(__file__).parent
TABLE_DIR = ROOT_FILE / "planilha" / "planilha_exemplo.xlsx"

# Carrega a planilha existente
workbook = load_workbook(TABLE_DIR)
# Nome para a planilha
new_sheet_name = 'Codigo supressao'
# Verifica se a aba já existe
if new_sheet_name not in workbook.sheetnames:
    # Adiciona a nova aba
    new_worksheet = workbook.create_sheet(new_sheet_name)
else:
    # Seleciona a aba existente
    new_worksheet = workbook[new_sheet_name]

# Seleciona a aba existente (Aba 1)
origin_sheet_name = 'Codigos Car'
origin_worksheet = workbook[origin_sheet_name]

# criando index com base em colunas
first_column = origin_worksheet.min_column
list_colums = []
for i in range(first_column,first_column+6):
    list_colums.append(i)

# # Adicionando uma nova coluna após a última coluna com valores
# new_worksheet.insert_cols(1)

list_names = []
dict_car = { }

# Concertando as celulas da worksheet original
row:tuple[Cell]
for row in origin_worksheet.iter_rows():
    for cell in row:
        value = str(cell.value)
        value = value.replace(",","")
        # Substituir &lt; por <
        value = value.replace("&amp;lt;", "<")
        value = value.replace("&lt;", "<")
        value = value.replace("&amp;", "<")
        value = value.replace("< ","<")
        cell.number_format = "General"
        cell.value = '' if cell.value is None else value
        cell.value = '' if value == "None" else value

        if cell.column == list_colums[2]:
            if cell.value is not None and cell.value != "" and cell.value != "None":
                if cell.value not in dict_car:
                    dict_car[cell.value] = 1
                else:
                    dict_car[cell.value] += 1

row: tuple[Cell]
for row in origin_worksheet.iter_rows(min_row=2):
    line = ""
    for cell in row:
        if cell.value is not None:

            car_column = cell.value

            if cell.column == list_colums[0]:
                words = str(cell.value).split()
                first_chars = [word[0] for word in words]
                line += "".join(first_chars)
                line += "".join("-")

            elif cell.column == list_colums[1]:
                
                number = str(cell.value).zfill(4)
                line += number
                line += "".join("-")

            elif cell.column == list_colums[2]:
                if car_column in dict_car:
                    count_repeat = dict_car[car_column]
                    count_repeat = str(count_repeat).zfill(2)
                    line += "".join(count_repeat)
                    line += "".join("-")
                else:
                    continue

            elif cell.column == list_colums[3]:
                words = str(cell.value).split()
                first_chars = [word[0] for word in words]
                line += "".join(first_chars)
            
            elif cell.column == list_colums[4]:
                new_value = str(cell.value)
                new_value = new_value.replace(" ","")
                new_value = new_value.replace(",","")
                new_value = new_value.upper()
                if new_value == "SEMCULTIVO":
                    line += "000"
                else:
                    anos = new_value.split("/")
                    ultimos_digito_por_ano = [ano[-1] if len(ano) > 0 else '0' for ano in anos]
                    line += "".join(ultimos_digito_por_ano).zfill(3)
                line += "".join("-")

            elif cell.column == list_colums[5]:
                cell.value = str(cell.value).replace("&lt;","<")
                if cell.value == "<2017":
                    line += "".join("01")
                else:
                    year = str(cell.value)[-2:]
                    line += "".join(year)
            
    if line:
        list_names.append(line)

# Adicione os valores nas células da nova coluna (supondo que list_names contenha os dados)
for index, value in enumerate(list_names, start=2):
    new_worksheet.cell(row=index, column=1, value=value)

#corrigindo formatações
row:tuple[Cell]
for row in new_worksheet.iter_rows():
    for cell in row:
        value = str(cell.value)
        value = value.replace("&amp;lt;", "<")
        value = value.replace("&lt;", "<")
        value = value.replace("&amp;", "<")
        # if value == "-0000-000-":
        #     value = ""
        cell.value = value
        if cell.row == 1 and cell.column == 1:
            cell.value = "COD SUP"

workbook.save(TABLE_DIR)
