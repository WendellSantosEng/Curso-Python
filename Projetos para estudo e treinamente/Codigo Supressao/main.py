from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FILE = Path(__file__).parent
TABLE_DIR = ROOT_FILE / "planilha" / "planilhateste.xlsx"

workbook = Workbook()
workbook:Workbook = load_workbook(TABLE_DIR)

sheet_name = "Codigos Car"
worksheet: Worksheet = workbook[sheet_name]

# Encontre o índice da última coluna com valores
last_column_index = max(worksheet.max_column, 1)  # Certifique-se de que o índice não seja inferior a 1

# Adicione uma nova coluna após a última coluna com valores
worksheet.insert_cols(last_column_index + 1)

list_names = []

row: tuple[Cell]
for row in worksheet.iter_rows():
    line = ""
    for cell in row:
        if cell.value is not None:
            words = cell.value.split()
            first_chars = [word[0] for word in words]
            line += "".join(first_chars)
    if line:
        list_names.append(line)

# Adicione os valores nas células da nova coluna (supondo que list_names contenha os dados)
for index, value in enumerate(list_names, start=3):
    worksheet.cell(row=index, column=last_column_index + 1, value=value)


workbook.save(TABLE_DIR)
