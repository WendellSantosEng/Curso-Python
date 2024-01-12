from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FILE = Path(__file__).parent
WORKBOOK_PATH = ROOT_FILE / "planilhas" / "planilha.xlsx"

workbook:Workbook = load_workbook(WORKBOOK_PATH)

# worksheet:Worksheet = workbook.active   cria uma aba de planilha e ativa ela pra fazer mudança

# Nome para a planilha
sheet_name = 'Minha planilha'
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end="\t")

        if cell.value == "Alberto":
            worksheet.cell(cell.row,1,"Felipe")

    print()

worksheet["A3"].value = "Otavio"

workbook.save(WORKBOOK_PATH)

