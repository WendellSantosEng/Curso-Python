from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FILE = Path(__file__).parent
WORKBOOK_PATH = ROOT_FILE / "planilhas" / "planilha.xlsx"

workbook = Workbook()

# worksheet:Worksheet = workbook.active   cria uma aba de planilha e ativa ela pra fazer mudança

# Nome para a planilha
sheet_name = 'Minha planilha'
# Criamos a planilha
workbook.create_sheet(sheet_name, 0) # o indice 0 é para mover ela pro inicio, opcional
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])


worksheet.cell(1,1, "NOME")
worksheet.cell(1,2, "IDADE")
worksheet.cell(1,3, "NOTA")

students = [
    # nome      idade nota
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, students_row in enumerate(students, start=3):
#     for j, students_columm in enumerate(students_row, start=1):
#         print(i,j,students_columm)
#         worksheet.cell(i,j,students_columm)

for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)

